"""Parse a historical-invoices bulk file (Excel or CSV).

Two sheets / two files are supported:

- `invoices` — one row per historical invoice.
- `mappings` — one row per (invoice_number, boq_line_number) mapping, with quantity
  and amount billed against that BoQ line.

If a single-sheet CSV is provided it is treated as the `invoices` sheet and no
BoQ mappings are attached (caller may supply a second CSV for mappings).
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from datetime import date
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

INVOICE_REQUIRED = {"invoice_number", "vendor_trn", "contract_number", "invoice_date", "total"}

_INVOICE_ALIASES = {
    "invoice no": "invoice_number",
    "invoice #": "invoice_number",
    "vendor trn": "vendor_trn",
    "trn": "vendor_trn",
    "vat number": "vendor_trn",
    "contract": "contract_number",
    "contract #": "contract_number",
    "contract no": "contract_number",
    "date": "invoice_date",
    "subtotal": "subtotal",
    "vat": "vat",
    "total": "total",
    "status": "status",
    "paid amount": "paid_amount",
    "payment date": "payment_date",
    "payment reference": "payment_reference",
    "payment ref": "payment_reference",
}

_MAPPING_ALIASES = {
    "invoice no": "invoice_number",
    "invoice #": "invoice_number",
    "boq line": "boq_line_number",
    "line": "boq_line_number",
    "line no": "boq_line_number",
    "line #": "boq_line_number",
    "qty": "quantity",
    "quantity": "quantity",
    "amount": "amount",
    "line total": "amount",
    "total": "amount",
}


@dataclass
class HistoricalInvoiceRow:
    invoice_number: str
    vendor_trn: str
    contract_number: str
    invoice_date: date
    subtotal: Decimal
    vat: Decimal
    total: Decimal
    status: str = "unpaid"
    paid_amount: Decimal = Decimal("0")
    payment_date: date | None = None
    payment_reference: str | None = None
    errors: list[str] = field(default_factory=list)


@dataclass
class HistoricalMappingRow:
    invoice_number: str
    boq_line_number: int
    quantity: Decimal
    amount: Decimal
    errors: list[str] = field(default_factory=list)


@dataclass
class HistoricalPreview:
    invoices: list[HistoricalInvoiceRow]
    mappings: list[HistoricalMappingRow]
    row_errors: int


def _norm(h: str, aliases: dict[str, str]) -> str:
    key = h.strip().lower()
    if key in aliases:
        return aliases[key]
    return key.replace(" ", "_")


def _to_decimal(v: object) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    return Decimal(str(v).strip().replace(",", ""))


def _to_date(v: object) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, date):
        return v
    s = str(v).strip()
    # accept ISO and DD/MM/YYYY
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            from datetime import datetime

            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"invalid date: {v!r}")


def _parse_invoice_row(row: dict[str, object]) -> HistoricalInvoiceRow:
    errors: list[str] = []
    try:
        inv_date = _to_date(row.get("invoice_date"))
    except ValueError as exc:
        inv_date = date.today()
        errors.append(str(exc))
    if inv_date is None:
        inv_date = date.today()
        errors.append("missing invoice_date")
    try:
        subtotal = _to_decimal(row.get("subtotal")).quantize(Decimal("0.01"))
        vat = _to_decimal(row.get("vat")).quantize(Decimal("0.01"))
        total = _to_decimal(row.get("total")).quantize(Decimal("0.01"))
        if total == 0 and (subtotal > 0 or vat > 0):
            total = (subtotal + vat).quantize(Decimal("0.01"))
        if subtotal == 0 and total > 0 and vat == 0:
            subtotal = total
    except InvalidOperation:
        subtotal = Decimal("0")
        vat = Decimal("0")
        total = Decimal("0")
        errors.append("invalid subtotal/vat/total")
    try:
        paid_amount = _to_decimal(row.get("paid_amount")).quantize(Decimal("0.01"))
    except InvalidOperation:
        paid_amount = Decimal("0")
        errors.append("invalid paid_amount")
    status = str(row.get("status") or ("paid" if paid_amount >= total and total > 0 else "unpaid")).strip().lower()
    try:
        payment_date = _to_date(row.get("payment_date"))
    except ValueError:
        payment_date = None
        errors.append("invalid payment_date")
    invoice_number = str(row.get("invoice_number") or "").strip()
    if not invoice_number:
        errors.append("invoice_number is required")
    vendor_trn = str(row.get("vendor_trn") or "").strip()
    contract_number = str(row.get("contract_number") or "").strip()
    if not contract_number:
        errors.append("contract_number is required")
    return HistoricalInvoiceRow(
        invoice_number=invoice_number,
        vendor_trn=vendor_trn,
        contract_number=contract_number,
        invoice_date=inv_date,
        subtotal=subtotal,
        vat=vat,
        total=total,
        status=status,
        paid_amount=paid_amount,
        payment_date=payment_date,
        payment_reference=str(row.get("payment_reference") or "").strip() or None,
        errors=errors,
    )


def _parse_mapping_row(row: dict[str, object]) -> HistoricalMappingRow:
    errors: list[str] = []
    try:
        line = int(_to_decimal(row.get("boq_line_number")))
    except (InvalidOperation, TypeError, ValueError):
        line = 0
        errors.append("invalid boq_line_number")
    try:
        qty = _to_decimal(row.get("quantity")).quantize(Decimal("0.0001"))
    except InvalidOperation:
        qty = Decimal("0")
        errors.append("invalid quantity")
    try:
        amount = _to_decimal(row.get("amount")).quantize(Decimal("0.01"))
    except InvalidOperation:
        amount = Decimal("0")
        errors.append("invalid amount")
    invoice_number = str(row.get("invoice_number") or "").strip()
    if not invoice_number:
        errors.append("invoice_number is required")
    return HistoricalMappingRow(
        invoice_number=invoice_number,
        boq_line_number=line,
        quantity=qty,
        amount=amount,
        errors=errors,
    )


def _read_csv(blob: bytes, aliases: dict[str, str]) -> list[dict[str, object]]:
    text = blob.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    try:
        header = [_norm(h, aliases) for h in next(reader)]
    except StopIteration:
        return []
    rows: list[dict[str, object]] = []
    for r in reader:
        if not any(cell.strip() for cell in r):
            continue
        rows.append(dict(zip(header, r, strict=False)))
    return rows


def _read_xlsx(
    blob: bytes,
) -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    wb = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    sheet_names = [s.lower() for s in wb.sheetnames]
    inv_sheet = wb[wb.sheetnames[sheet_names.index("invoices")]] if "invoices" in sheet_names else wb.worksheets[0]
    map_sheet = (
        wb[wb.sheetnames[sheet_names.index("mappings")]] if "mappings" in sheet_names else None
    )

    def read_sheet(ws, aliases: dict[str, str]) -> list[dict[str, object]]:
        it = ws.iter_rows(values_only=True)
        try:
            header = [_norm(str(h or ""), aliases) for h in next(it)]
        except StopIteration:
            return []
        out = []
        for r in it:
            if not any(c not in (None, "") for c in r):
                continue
            out.append(dict(zip(header, r, strict=False)))
        return out

    invoices = read_sheet(inv_sheet, _INVOICE_ALIASES)
    mappings = read_sheet(map_sheet, _MAPPING_ALIASES) if map_sheet else []
    return invoices, mappings


def parse_historical(
    filename: str,
    blob: bytes,
    mappings_blob: bytes | None = None,
    mappings_filename: str | None = None,
) -> HistoricalPreview:
    name = filename.lower()
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        inv_rows, map_rows = _read_xlsx(blob)
    elif name.endswith(".csv") or name.endswith(".tsv"):
        inv_rows = _read_csv(blob, _INVOICE_ALIASES)
        map_rows = (
            _read_csv(mappings_blob, _MAPPING_ALIASES)
            if mappings_blob and (mappings_filename or "").lower().endswith((".csv", ".tsv"))
            else []
        )
    else:
        raise ValueError(f"Unsupported file type: {filename}")

    missing_inv = INVOICE_REQUIRED - (set(inv_rows[0].keys()) if inv_rows else set())
    if inv_rows and missing_inv:
        raise ValueError(f"Invoices sheet missing columns: {sorted(missing_inv)}")

    invoices = [_parse_invoice_row(r) for r in inv_rows]
    mappings = [_parse_mapping_row(r) for r in map_rows]
    return HistoricalPreview(
        invoices=invoices,
        mappings=mappings,
        row_errors=sum(1 for r in invoices if r.errors) + sum(1 for r in mappings if r.errors),
    )
