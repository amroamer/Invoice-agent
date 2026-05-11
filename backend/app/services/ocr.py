"""Run OCR over invoice artifacts and return plain text.

Strategy:
 - PDFs: try pdfplumber text extraction first (fast, works for digitally-generated
   PDFs). If the result is empty or tiny, render pages via PyMuPDF and run
   Tesseract on each page image.
 - Images: Tesseract directly.
 - Excel: openpyxl values-only flatten (not strictly OCR, but useful for
   .xlsx "invoices").

The function never raises on missing optional deps — it logs and returns a best
effort. Callers should treat empty output as a signal to ask the officer to
re-upload or enter fields manually.
"""
from __future__ import annotations

import io
import logging
from dataclasses import dataclass

log = logging.getLogger(__name__)


@dataclass
class OcrResult:
    text: str
    page_count: int
    method: str


def _pdf_text_pdfplumber(blob: bytes) -> tuple[str, int]:
    try:
        import pdfplumber
    except ImportError:
        return "", 0
    chunks: list[str] = []
    pages = 0
    with pdfplumber.open(io.BytesIO(blob)) as pdf:
        pages = len(pdf.pages)
        for p in pdf.pages:
            t = p.extract_text() or ""
            if t.strip():
                chunks.append(t)
    return "\n\n".join(chunks).strip(), pages


def _pdf_text_tesseract(blob: bytes) -> tuple[str, int]:
    try:
        import fitz  # PyMuPDF
        import pytesseract
        from PIL import Image
    except ImportError as exc:
        log.warning("ocr_deps_missing", extra={"err": str(exc)})
        return "", 0
    chunks: list[str] = []
    pages = 0
    with fitz.open(stream=blob, filetype="pdf") as doc:
        pages = len(doc)
        for page in doc:
            pix = page.get_pixmap(dpi=250)
            img = Image.open(io.BytesIO(pix.tobytes("png")))
            chunks.append(pytesseract.image_to_string(img, lang="eng"))
    return "\n\n".join(chunks).strip(), pages


def _image_text(blob: bytes) -> str:
    try:
        import pytesseract
        from PIL import Image
    except ImportError as exc:
        log.warning("ocr_deps_missing", extra={"err": str(exc)})
        return ""
    img = Image.open(io.BytesIO(blob))
    return pytesseract.image_to_string(img, lang="eng").strip()


def _xlsx_text(blob: bytes) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ""
    wb = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    lines: list[str] = []
    for ws in wb.worksheets:
        lines.append(f"# {ws.title}")
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                lines.append("\t".join(cells))
    return "\n".join(lines).strip()


def run_ocr(blob: bytes, extension: str) -> OcrResult:
    ext = extension.lower().lstrip(".")
    if ext == "pdf":
        text, pages = _pdf_text_pdfplumber(blob)
        if len(text) >= 40:
            return OcrResult(text=text, page_count=pages, method="pdfplumber")
        log.info("pdf_falling_back_to_ocr", extra={"chars": len(text), "pages": pages})
        text, pages = _pdf_text_tesseract(blob)
        return OcrResult(text=text, page_count=pages or 1, method="tesseract")
    if ext in {"png", "jpg", "jpeg", "tiff"}:
        return OcrResult(text=_image_text(blob), page_count=1, method="tesseract")
    if ext in {"xlsx", "xlsm"}:
        return OcrResult(text=_xlsx_text(blob), page_count=1, method="xlsx")
    return OcrResult(text="", page_count=0, method="unsupported")
