"""Parse an uploaded BoQ file (Excel or CSV) into normalized line rows.

Parser is tolerant about header casing/whitespace/aliases but strict about the
shape it returns. Unknown columns are ignored; missing required columns raise.
"""
from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook

REQUIRED_COLS = {"line_number", "description", "uom", "quantity", "unit_price"}

_ALIASES: dict[str, str] = {
    "line": "line_number",
    "line no": "line_number",
    "line #": "line_number",
    "no": "line_number",
    "item": "line_number",
    "desc": "description",
    "description": "description",
    "unit": "uom",
    "unit of measure": "uom",
    "uom": "uom",
    "qty": "quantity",
    "quantity": "quantity",
    "unit price": "unit_price",
    "rate": "unit_price",
    "price": "unit_price",
    "total": "line_total",
    "amount": "line_total",
    "line total": "line_total",
}


@dataclass
class BoqRow:
    line_number: int
    description: str
    uom: str
    quantity: Decimal
    unit_price: Decimal
    line_total: Decimal
    errors: list[str] = field(default_factory=list)


@dataclass
class BoqPreview:
    rows: list[BoqRow]
    row_errors: int
    sum_line_total: Decimal
    contract_value: Decimal | None
    tolerance_pct: float
    within_tolerance: bool | None


def _normalize_header(h: str) -> str:
    return _ALIASES.get(h.strip().lower(), h.strip().lower().replace(" ", "_"))


def _to_decimal(v: object) -> Decimal:
    if v is None or v == "":
        raise InvalidOperation("empty")
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    s = str(v).strip().replace(",", "")
    return Decimal(s)


def _parse_rows(raw: list[dict[str, object]]) -> list[BoqRow]:
    parsed: list[BoqRow] = []
    for idx, row in enumerate(raw, start=1):
        errors: list[str] = []
        try:
            line_number = int(_to_decimal(row.get("line_number")))
        except (InvalidOperation, TypeError, ValueError):
            line_number = idx
            errors.append("invalid line_number, defaulted to row index")
        description = str(row.get("description") or "").strip()
        if not description:
            errors.append("description is required")
        uom = str(row.get("uom") or "").strip() or "unit"
        try:
            qty = _to_decimal(row.get("quantity"))
        except (InvalidOperation, TypeError, ValueError):
            qty = Decimal("0")
            errors.append("invalid quantity")
        try:
            price = _to_decimal(row.get("unit_price"))
        except (InvalidOperation, TypeError, ValueError):
            price = Decimal("0")
            errors.append("invalid unit_price")

        explicit_total = row.get("line_total")
        if explicit_total in (None, ""):
            line_total = (qty * price).quantize(Decimal("0.01"))
        else:
            try:
                line_total = _to_decimal(explicit_total).quantize(Decimal("0.01"))
                computed = (qty * price).quantize(Decimal("0.01"))
                if abs(line_total - computed) > Decimal("0.05"):
                    errors.append(
                        f"line_total ({line_total}) differs from qty*unit_price ({computed})"
                    )
            except (InvalidOperation, TypeError, ValueError):
                line_total = (qty * price).quantize(Decimal("0.01"))
                errors.append("invalid line_total, recomputed")

        parsed.append(
            BoqRow(
                line_number=line_number,
                description=description,
                uom=uom,
                quantity=qty,
                unit_price=price,
                line_total=line_total,
                errors=errors,
            )
        )
    return parsed


def _csv_to_rows(blob: bytes) -> list[dict[str, object]]:
    text = blob.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(text))
    try:
        header = [_normalize_header(h) for h in next(reader)]
    except StopIteration:
        return []
    missing = REQUIRED_COLS - set(header)
    if missing:
        raise ValueError(f"Missing required columns: {sorted(missing)}")
    rows: list[dict[str, object]] = []
    for r in reader:
        if not any(cell.strip() for cell in r):
            continue
        rows.append(dict(zip(header, r, strict=False)))
    return rows


def _xlsx_to_rows(blob: bytes) -> list[dict[str, object]]:
    wb = load_workbook(io.BytesIO(blob), data_only=True, read_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    try:
        header_row = next(it)
    except StopIteration:
        return []
    header = [_normalize_header(str(h or "")) for h in header_row]
    missing = REQUIRED_COLS - set(header)
    if missing:
        raise ValueError(f"Missing required columns: {sorted(missing)}")
    rows: list[dict[str, object]] = []
    for r in it:
        if not any(c not in (None, "") for c in r):
            continue
        rows.append(dict(zip(header, r, strict=False)))
    return rows


def parse_boq(filename: str, blob: bytes) -> list[BoqRow]:
    name = filename.lower()
    if name.endswith(".csv") or name.endswith(".tsv"):
        return _parse_rows(_csv_to_rows(blob))
    if name.endswith(".xlsx") or name.endswith(".xlsm"):
        return _parse_rows(_xlsx_to_rows(blob))
    raise ValueError(f"Unsupported BoQ file type: {filename}")


def preview(
    rows: list[BoqRow],
    *,
    contract_value: Decimal | None,
    tolerance_pct: float,
) -> BoqPreview:
    total = sum((r.line_total for r in rows), Decimal("0"))
    if contract_value is None or contract_value == 0:
        within = None
    else:
        diff_pct = float(abs(total - contract_value) / contract_value * 100)
        within = diff_pct <= tolerance_pct
    return BoqPreview(
        rows=rows,
        row_errors=sum(1 for r in rows if r.errors),
        sum_line_total=total.quantize(Decimal("0.01")),
        contract_value=contract_value,
        tolerance_pct=tolerance_pct,
        within_tolerance=within,
    )
