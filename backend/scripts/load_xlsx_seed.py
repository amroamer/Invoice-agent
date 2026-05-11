"""Load the Finance_Invoicing_Agent_Seed_Data.xlsx workbook into the app DB.

Idempotent:
- upserts users by username
- upserts projects by (name, client_entity)
- upserts vendors by trn
- upserts contracts by (vendor, contract_number)
- replaces active BoQ rows per contract (archives the old ones)
- upserts historical invoices by (vendor, invoice_number) — re-attaches line items
  and the single payment row if any

Run:
    docker compose exec -T backend python /app/scripts/load_xlsx_seed.py /tmp/seed.xlsx
"""
from __future__ import annotations

import re
import sys
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from app.core.security import hash_password
from app.db.session import SessionLocal
from app.models.boq_item import BoqItem
from app.models.contract import Contract
from app.models.enums import (
    InvoiceSource,
    InvoiceStatus,
    ProjectStatus,
    UserRole,
    VatTreatment,
)
from app.models.invoice import Invoice
from app.models.invoice_line_item import InvoiceLineItem
from app.models.payment import Payment
from app.models.project import Project
from app.models.user import User
from app.models.vendor import Vendor


STATUS_MAP = {
    "paid": InvoiceStatus.paid,
    "unpaid": InvoiceStatus.pending,
    "pending": InvoiceStatus.pending,
    "partial": InvoiceStatus.partially_paid,
    "partially_paid": InvoiceStatus.partially_paid,
}

PROJECT_STATUS_MAP = {
    "active": ProjectStatus.active,
    "on_hold": ProjectStatus.on_hold,
    "closed": ProjectStatus.closed,
    "inactive": ProjectStatus.inactive,
}

VAT_TREATMENT_MAP = {
    "15% standard": VatTreatment.exclusive,
    "standard": VatTreatment.exclusive,
    "exclusive": VatTreatment.exclusive,
    "inclusive": VatTreatment.inclusive,
    "exempt": VatTreatment.exempt,
}


def _to_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    return s in {"true", "1", "yes", "y"}


def _to_date(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"bad date: {v!r}")


def _to_dec(v, default="0") -> Decimal:
    if v is None or v == "":
        return Decimal(default)
    if isinstance(v, Decimal):
        return v
    if isinstance(v, (int, float)):
        return Decimal(str(v))
    return Decimal(str(v).strip().replace(",", ""))


def _to_int_line(v) -> int:
    """'L001' → 1, '3' → 3."""
    if v is None:
        return 0
    s = str(v).strip()
    m = re.match(r"^[A-Za-z]*0*(\d+)$", s)
    if m:
        return int(m.group(1))
    return int(float(s))


def _sheet_rows(wb, name: str) -> list[dict]:
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    out: list[dict] = []
    for r in rows[1:]:
        if all(c is None or c == "" for c in r):
            continue
        out.append(dict(zip(headers, r, strict=False)))
    return out


def load_users(db, rows) -> int:
    touched = 0
    for r in rows:
        username = str(r.get("username") or "").strip()
        if not username:
            continue
        user = db.scalar(select(User).where(User.username == username))
        role = UserRole(str(r.get("role") or "officer").strip().lower())
        active = _to_bool(r.get("active", True))
        pwd = str(r.get("initial_password") or "").strip()
        email = str(r.get("email") or "").strip()
        full_name = str(r.get("full_name") or "").strip() or None

        if user is None:
            user = User(
                username=username,
                email=email,
                full_name=full_name,
                password_hash=hash_password(pwd or "ChangeMe!123"),
                role=role,
                active=active,
            )
            db.add(user)
        else:
            user.email = email or user.email
            user.full_name = full_name or user.full_name
            user.role = role
            user.active = active
            if pwd:
                user.password_hash = hash_password(pwd)
        touched += 1
    db.flush()
    return touched


def load_projects(db, rows, admin_id) -> dict[str, Project]:
    out: dict[str, Project] = {}
    for r in rows:
        code = str(r.get("project_code") or "").strip()
        name = str(r.get("name") or "").strip()
        client = str(r.get("client_entity") or "").strip()
        if not (code and name and client):
            continue
        project = db.scalar(
            select(Project).where(Project.name == name, Project.client_entity == client)
        )
        status = PROJECT_STATUS_MAP.get(
            str(r.get("status") or "active").strip().lower(), ProjectStatus.active
        )
        if project is None:
            project = Project(
                name=name,
                client_entity=client,
                description=str(r.get("description") or "") or None,
                start_date=_to_date(r.get("start_date")),
                end_date=_to_date(r.get("end_date")),
                status=status,
                created_by=admin_id,
                updated_by=admin_id,
            )
            db.add(project)
        else:
            project.description = str(r.get("description") or "") or project.description
            project.start_date = _to_date(r.get("start_date")) or project.start_date
            project.end_date = _to_date(r.get("end_date")) or project.end_date
            project.status = status
            project.updated_by = admin_id
        db.flush()
        out[code] = project
    return out


def load_vendors(db, rows, admin_id) -> dict[str, Vendor]:
    out: dict[str, Vendor] = {}
    for r in rows:
        code = str(r.get("vendor_code") or "").strip()
        trn = str(r.get("trn") or "").strip()
        if not (code and trn):
            continue
        vendor = db.scalar(select(Vendor).where(Vendor.trn == trn))
        bank_iban = str(r.get("bank_iban") or "").strip()
        bank_details = {"iban": bank_iban} if bank_iban else None
        if vendor is None:
            vendor = Vendor(
                legal_name=str(r.get("legal_name") or "").strip(),
                trn=trn,
                cr_number=str(r.get("cr_number") or "").strip() or None,
                contact_email=str(r.get("contact_email") or "").strip() or None,
                bank_details=bank_details,
                active=_to_bool(r.get("active", True)),
                created_by=admin_id,
                updated_by=admin_id,
            )
            db.add(vendor)
        else:
            vendor.legal_name = str(r.get("legal_name") or vendor.legal_name).strip()
            vendor.cr_number = str(r.get("cr_number") or "").strip() or vendor.cr_number
            vendor.contact_email = (
                str(r.get("contact_email") or "").strip() or vendor.contact_email
            )
            if bank_details:
                vendor.bank_details = bank_details
            vendor.active = _to_bool(r.get("active", True))
            vendor.updated_by = admin_id
        db.flush()
        out[code] = vendor
    return out


def load_contracts(
    db, rows, projects: dict[str, Project], vendors: dict[str, Vendor], admin_id
) -> dict[str, Contract]:
    out: dict[str, Contract] = {}
    for r in rows:
        number = str(r.get("contract_number") or "").strip()
        pcode = str(r.get("project_code") or "").strip()
        vcode = str(r.get("vendor_code") or "").strip()
        if not (number and pcode in projects and vcode in vendors):
            continue
        vendor = vendors[vcode]
        contract = db.scalar(
            select(Contract).where(
                Contract.vendor_id == vendor.id, Contract.contract_number == number
            )
        )
        vat = VAT_TREATMENT_MAP.get(
            str(r.get("vat_treatment") or "").strip().lower(), VatTreatment.exclusive
        )
        status = PROJECT_STATUS_MAP.get(
            str(r.get("status") or "active").strip().lower(), ProjectStatus.active
        )
        if contract is None:
            contract = Contract(
                project_id=projects[pcode].id,
                vendor_id=vendor.id,
                contract_number=number,
                value=_to_dec(r.get("value_sar")),
                currency=str(r.get("currency") or "SAR").strip()[:3],
                start_date=_to_date(r.get("start_date")) or date.today(),
                end_date=_to_date(r.get("end_date")) or date.today(),
                retention_pct=_to_dec(r.get("retention_pct")),
                advance_pct=_to_dec(r.get("advance_pct")),
                vat_treatment=vat,
                status=status,
                created_by=admin_id,
                updated_by=admin_id,
            )
            db.add(contract)
        else:
            contract.project_id = projects[pcode].id
            contract.value = _to_dec(r.get("value_sar"))
            contract.start_date = _to_date(r.get("start_date")) or contract.start_date
            contract.end_date = _to_date(r.get("end_date")) or contract.end_date
            contract.retention_pct = _to_dec(r.get("retention_pct"))
            contract.advance_pct = _to_dec(r.get("advance_pct"))
            contract.vat_treatment = vat
            contract.status = status
            contract.updated_by = admin_id
        db.flush()
        out[number] = contract
    return out


def load_boq(
    db, rows, contracts: dict[str, Contract], admin_id
) -> dict[tuple[str, int], BoqItem]:
    """Replace active BoQ per contract: archive old, insert new."""
    # Deactivate existing active BoQ for contracts we're touching.
    touched_contracts = {
        str(r.get("contract_number") or "").strip() for r in rows
    } & set(contracts.keys())
    for cn in touched_contracts:
        for prior in db.scalars(
            select(BoqItem).where(
                BoqItem.contract_id == contracts[cn].id, BoqItem.active.is_(True)
            )
        ):
            prior.active = False
    db.flush()

    out: dict[tuple[str, int], BoqItem] = {}
    for r in rows:
        cn = str(r.get("contract_number") or "").strip()
        if cn not in contracts:
            continue
        line_no = _to_int_line(r.get("line_number"))
        description = str(r.get("description") or "").strip()
        uom = str(r.get("uom") or "unit").strip()
        qty = _to_dec(r.get("quantity"))
        unit_price = _to_dec(r.get("unit_price_sar"))
        line_total = _to_dec(r.get("line_total_sar"))
        if line_total == 0:
            line_total = (qty * unit_price).quantize(Decimal("0.01"))
        item = BoqItem(
            contract_id=contracts[cn].id,
            line_number=line_no,
            description=description,
            uom=uom,
            quantity=qty,
            unit_price=unit_price,
            line_total=line_total,
            active=True,
            created_by=admin_id,
            updated_by=admin_id,
        )
        db.add(item)
        db.flush()
        out[(cn, line_no)] = item
    return out


def load_historical_invoices(
    db,
    rows,
    contracts: dict[str, Contract],
    vendors: dict[str, Vendor],
    admin_id,
) -> dict[str, Invoice]:
    out: dict[str, Invoice] = {}
    for r in rows:
        number = str(r.get("invoice_number") or "").strip()
        cn = str(r.get("contract_number") or "").strip()
        vcode = str(r.get("vendor_code") or "").strip()
        if not (number and cn in contracts and vcode in vendors):
            continue
        contract = contracts[cn]
        vendor = vendors[vcode]
        status = STATUS_MAP.get(
            str(r.get("status") or "unpaid").strip().lower(), InvoiceStatus.pending
        )
        subtotal = _to_dec(r.get("subtotal_sar"))
        vat = _to_dec(r.get("vat_sar"))
        total = _to_dec(r.get("total_sar"))
        if total == 0 and subtotal + vat > 0:
            total = (subtotal + vat).quantize(Decimal("0.01"))
        invoice_date = _to_date(r.get("invoice_date")) or date.today()

        invoice = db.scalar(
            select(Invoice).where(
                Invoice.vendor_id == vendor.id,
                Invoice.invoice_number == number,
                Invoice.archived.is_(False),
            )
        )
        if invoice is None:
            invoice = Invoice(
                contract_id=contract.id,
                vendor_id=vendor.id,
                invoice_number=number,
                invoice_date=invoice_date,
                subtotal=subtotal,
                vat=vat,
                total=total,
                currency=contract.currency,
                source=InvoiceSource.historical,
                status=status,
                uploaded_by=admin_id,
                created_by=admin_id,
                updated_by=admin_id,
            )
            db.add(invoice)
            db.flush()
        else:
            invoice.contract_id = contract.id
            invoice.invoice_date = invoice_date
            invoice.subtotal = subtotal
            invoice.vat = vat
            invoice.total = total
            invoice.status = status
            invoice.updated_by = admin_id
            # clear prior lines+payments so we can re-attach cleanly
            for line in list(invoice.line_items):
                db.delete(line)
            for p in db.scalars(select(Payment).where(Payment.invoice_id == invoice.id)):
                db.delete(p)
            db.flush()

        paid_amount = _to_dec(r.get("paid_amount_sar"))
        if paid_amount > 0:
            db.add(
                Payment(
                    invoice_id=invoice.id,
                    amount=paid_amount,
                    payment_date=_to_date(r.get("payment_date")) or invoice_date,
                    reference=str(r.get("payment_reference") or f"HIST-{number}").strip(),
                    recorded_by=admin_id,
                )
            )

        out[number] = invoice
    db.flush()
    return out


def load_historical_lines(
    db,
    rows,
    invoices: dict[str, Invoice],
    boq_by_contract_line: dict[tuple[str, int], BoqItem],
) -> int:
    # Map invoice → contract_number for BoQ lookups
    inv_to_contract: dict[str, str] = {}
    for inv_number, inv in invoices.items():
        contract = inv.contract_id and db.get(Contract, inv.contract_id)
        if contract:
            inv_to_contract[inv_number] = contract.contract_number

    inserted = 0
    for r in rows:
        inv_number = str(r.get("invoice_number") or "").strip()
        if inv_number not in invoices:
            continue
        invoice = invoices[inv_number]
        line_no = _to_int_line(r.get("boq_line_number"))
        contract_number = inv_to_contract.get(inv_number)
        boq_item = boq_by_contract_line.get((contract_number, line_no)) if contract_number else None

        description = str(r.get("description") or "").strip()
        qty = _to_dec(r.get("quantity"))
        unit_price = _to_dec(r.get("unit_price_sar"))
        line_total = _to_dec(r.get("line_total_sar"))
        if line_total == 0 and (qty * unit_price) > 0:
            line_total = (qty * unit_price).quantize(Decimal("0.01"))

        db.add(
            InvoiceLineItem(
                invoice_id=invoice.id,
                boq_item_id=boq_item.id if boq_item else None,
                line_number=line_no,
                description=description or (boq_item.description if boq_item else ""),
                uom=boq_item.uom if boq_item else None,
                quantity=qty,
                unit_price=unit_price,
                line_total=line_total,
                not_in_boq=boq_item is None,
            )
        )
        inserted += 1
    db.flush()
    return inserted


def main(path: str) -> int:
    wb = load_workbook(Path(path), data_only=True)
    with SessionLocal() as db:
        # Need an admin user id for audit columns. Use first admin or any user.
        admin_rows = _sheet_rows(wb, "Users")
        touched_users = load_users(db, admin_rows)
        db.flush()
        admin = db.scalar(select(User).where(User.role == UserRole.admin))
        if admin is None:
            print("[seed] ERROR: no admin user after loading Users sheet")
            return 2
        admin_id = admin.id

        projects = load_projects(db, _sheet_rows(wb, "Projects"), admin_id)
        vendors = load_vendors(db, _sheet_rows(wb, "Vendors"), admin_id)
        contracts = load_contracts(
            db, _sheet_rows(wb, "Contracts"), projects, vendors, admin_id
        )
        boq = load_boq(db, _sheet_rows(wb, "BoQ"), contracts, admin_id)
        invoices = load_historical_invoices(
            db,
            _sheet_rows(wb, "Historical_Invoices"),
            contracts,
            vendors,
            admin_id,
        )
        line_count = load_historical_lines(
            db, _sheet_rows(wb, "Historical_Invoice_Lines"), invoices, boq
        )
        db.commit()

    print(
        f"[seed] users={touched_users} projects={len(projects)} vendors={len(vendors)} "
        f"contracts={len(contracts)} boq_lines={len(boq)} "
        f"historical_invoices={len(invoices)} invoice_line_items={line_count}"
    )
    return 0


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else "/tmp/seed.xlsx"
    sys.exit(main(src))
